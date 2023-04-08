import requests
from bs4 import BeautifulSoup
import openpyxl

# URL ---> https://www.imdb.com/chart/top

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='Title'
print(excel.sheetnames)
sheet.append(['Article Name','Author Name'])


try:
   source = requests.get('https://www.theverge.com/23650428/colorblindness-design-ui-accessibility-wordle')
   source.raise_for_status()

   soup = BeautifulSoup(source.text,'html.parser')

   movies = soup.find('tbody',class_='lister-list').find_all('tr')


   for mov in movies:
     name = mov.find('td',class_='titleColumn').a.text
     rank = mov.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]
     year = mov.find('td',class_='titleColumn').span.text.strip()
     element = mov.find('td',class_="duet--article--feature-headline  hover:shadow-underline-inherit").strong.text
     print(article_name,author_name)
     sheet.append([article_name,author_name])

except Exception as e:
    print(e)
  
excel.save('data.xlsx')